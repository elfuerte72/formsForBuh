"""1С export parser (SDK boundary: xlrd + openpyxl + stdlib csv).

This module is the ONLY place that imports ``xlrd``, ``openpyxl`` and
``csv``. SDK exceptions are translated into :class:`OneCParseError` so
pipelines keep handling a single error taxonomy.

The expected input is the standard 1С report
«Реестр документов "Поступление (акт, накладная, УПД)"» exported as
``.xls``, ``.xlsx`` or ``.csv``. Header row is auto-detected (it contains
the cells ``Дата``, ``Номер`` and ``Сумма``); columns are mapped by name
(``Номер вх.`` → ``upd_number``; ``Дата вх.`` or ``Дата`` → ``date``;
``Сумма`` → ``amount``; ``Информация`` → ``organization``). The ``Номер``
column (internal 1С document number) is intentionally ignored.
"""

from __future__ import annotations

import csv
import io
from datetime import date as Date
from datetime import datetime
from io import BytesIO
from typing import Any, Iterable

import openpyxl
import xlrd
from openpyxl.utils.exceptions import InvalidFileException

from app.core.errors import OneCParseError
from app.core.logging import get_logger
from app.models import OneCRecord

log = get_logger("onec")


_EXT_XLS = ".xls"
_EXT_XLSX = ".xlsx"
_EXT_CSV = ".csv"
_SUPPORTED_EXTS = (_EXT_XLS, _EXT_XLSX, _EXT_CSV)

_REQUIRED_HEADER_TOKENS = ("дата", "номер", "сумма")
_HEADER_SCAN_LIMIT = 30  # rows
_TOTAL_MARKERS = ("итого", "всего")
_SIGNATURE_MARKERS = ("ответственный", "(должность)", "(подпись)", "———", "—")

_NORMAL_KEY_UPD_NUMBER = ("номер вх.", "номер вх", "номер входящий", "входящий номер")
_NORMAL_KEY_DATE_PRIMARY = ("дата вх.", "дата вх", "дата входящая", "входящая дата")
_NORMAL_KEY_DATE_FALLBACK = ("дата",)
_NORMAL_KEY_AMOUNT = ("сумма",)
_NORMAL_KEY_ORG = ("информация", "контрагент")


class OneCParserService:
    """Stateless parser for the 1С documents register export."""

    def parse(self, data: bytes, filename: str) -> list[OneCRecord]:
        """Return :class:`OneCRecord` instances from ``data``.

        ``filename`` decides the strategy (extension whitelist).
        Raises :class:`OneCParseError` for any parser failure — including
        empty files and unsupported extensions.
        """
        ext = _resolve_ext(filename)
        log.info(
            "onec.parse.start",
            filename=filename,
            ext=ext,
            bytes=len(data),
        )

        if not data:
            raise OneCParseError("Файл пустой")

        try:
            if ext == _EXT_XLS:
                rows = _read_xls(data)
            elif ext == _EXT_XLSX:
                rows = _read_xlsx(data)
            else:
                rows = _read_csv(data)
        except OneCParseError:
            raise
        except Exception as exc:  # pragma: no cover - defensive
            log.exception("onec.parse.read_failed", filename=filename, ext=ext)
            raise OneCParseError(f"Не удалось прочитать файл: {exc}") from exc

        records = list(_extract_records(rows))
        log.info(
            "onec.parse.done",
            filename=filename,
            rows=len(rows),
            records=len(records),
        )
        return records


# --- file readers -----------------------------------------------------------


def _resolve_ext(filename: str) -> str:
    name = (filename or "").lower().strip()
    for ext in _SUPPORTED_EXTS:
        if name.endswith(ext):
            return ext
    raise OneCParseError(
        f"Неподдерживаемый формат: {filename!r} (ожидается .xls/.xlsx/.csv)"
    )


def _read_xls(data: bytes) -> list[list[Any]]:
    try:
        wb = xlrd.open_workbook(file_contents=data)
    except xlrd.XLRDError as exc:
        raise OneCParseError(f"xlrd не смог открыть .xls: {exc}") from exc
    if wb.nsheets == 0:
        raise OneCParseError("В .xls нет листов")
    sheet = wb.sheet_by_index(0)
    datemode = wb.datemode
    rows: list[list[Any]] = []
    for r in range(sheet.nrows):
        row: list[Any] = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    row.append(xlrd.xldate.xldate_as_datetime(cell.value, datemode))
                except (ValueError, xlrd.XLDateError):
                    row.append(cell.value)
            else:
                row.append(cell.value)
        rows.append(row)
    return rows


def _read_xlsx(data: bytes) -> list[list[Any]]:
    try:
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    except (InvalidFileException, OSError, KeyError, ValueError) as exc:
        raise OneCParseError(f"openpyxl не смог открыть .xlsx: {exc}") from exc
    if not wb.sheetnames:
        raise OneCParseError("В .xlsx нет листов")
    sheet = wb[wb.sheetnames[0]]
    rows: list[list[Any]] = []
    for raw in sheet.iter_rows(values_only=True):
        rows.append(list(raw))
    wb.close()
    return rows


def _read_csv(data: bytes) -> list[list[Any]]:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = data.decode("cp1251")
        except UnicodeDecodeError as exc:
            raise OneCParseError(f"Не удалось определить кодировку CSV: {exc}") from exc

    sample = text[:1024]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"

    reader = csv.reader(io.StringIO(text), dialect=dialect)
    return [list(r) for r in reader]


# --- record extraction ------------------------------------------------------


def _extract_records(rows: list[list[Any]]) -> Iterable[OneCRecord]:
    header_idx, mapping = _find_header(rows)
    log.debug("onec.parse.header_found", row=header_idx, mapping=mapping)

    for raw_idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if _is_total_row(row):
            log.debug("onec.parse.stop_total", row=raw_idx)
            break
        if _is_signature_row(row):
            log.debug("onec.parse.stop_signature", row=raw_idx)
            break
        if _is_blank_row(row):
            log.debug("onec.parse.skip_blank", row=raw_idx)
            continue

        try:
            record = _row_to_record(row, mapping=mapping, source_row=raw_idx)
        except _SkipRow as exc:
            log.warning(
                "onec.parse.skip_row",
                row=raw_idx,
                reason=exc.reason,
            )
            continue
        log.debug(
            "onec.parse.row",
            row=raw_idx,
            upd_number=record.upd_number,
            amount=record.amount,
        )
        yield record


def _find_header(rows: list[list[Any]]) -> tuple[int, dict[str, int]]:
    """Return ``(header_index, {field_name: column_index})``.

    Required: ``upd_number`` and one of date/amount must resolve. We allow
    fallback from ``Дата вх.`` to ``Дата``; ``Информация`` is optional.
    """
    for idx, row in enumerate(rows[:_HEADER_SCAN_LIMIT]):
        normalised = [_normalize_header_cell(c) for c in row]
        joined = " ".join(normalised)
        if not all(token in joined for token in _REQUIRED_HEADER_TOKENS):
            continue

        mapping: dict[str, int] = {}
        if (col := _match_column(normalised, _NORMAL_KEY_UPD_NUMBER)) is not None:
            mapping["upd_number"] = col
        if (col := _match_column(normalised, _NORMAL_KEY_DATE_PRIMARY)) is not None:
            mapping["date"] = col
        elif (col := _match_column(normalised, _NORMAL_KEY_DATE_FALLBACK)) is not None:
            mapping["date"] = col
        if (col := _match_column(normalised, _NORMAL_KEY_AMOUNT)) is not None:
            mapping["amount"] = col
        if (col := _match_column(normalised, _NORMAL_KEY_ORG)) is not None:
            mapping["organization"] = col

        if "upd_number" not in mapping:
            continue
        if "date" not in mapping or "amount" not in mapping:
            continue
        return idx, mapping

    raise OneCParseError(
        "Не удалось найти строку заголовка с колонками "
        "«Номер вх.», «Дата (вх.)» и «Сумма»."
    )


def _row_to_record(
    row: list[Any], *, mapping: dict[str, int], source_row: int
) -> OneCRecord:
    upd_raw = _cell(row, mapping.get("upd_number"))
    upd_number = _to_str(upd_raw)
    if not upd_number:
        raise _SkipRow("empty upd_number")

    date_value = _to_date(_cell(row, mapping.get("date")))
    amount_value = _to_amount(_cell(row, mapping.get("amount")))
    org_value = _to_str(_cell(row, mapping.get("organization"))) or None

    return OneCRecord(
        upd_number=upd_number,
        date=date_value,
        amount=amount_value,
        organization=org_value,
        source_row=source_row,
    )


# --- helpers ----------------------------------------------------------------


class _SkipRow(Exception):
    def __init__(self, reason: str) -> None:
        super().__init__(reason)
        self.reason = reason


def _normalize_header_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace("\xa0", " ")


def _match_column(normalised: list[str], candidates: tuple[str, ...]) -> int | None:
    for col, cell in enumerate(normalised):
        if not cell:
            continue
        for cand in candidates:
            if cell == cand or cand in cell:
                return col
    return None


def _cell(row: list[Any], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _is_blank_row(row: list[Any]) -> bool:
    return all(_cell_blank(c) for c in row)


def _cell_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _is_total_row(row: list[Any]) -> bool:
    for cell in row[:3]:
        if isinstance(cell, str) and cell.strip().lower().startswith(_TOTAL_MARKERS):
            return True
    return False


def _is_signature_row(row: list[Any]) -> bool:
    for cell in row:
        if isinstance(cell, str):
            text = cell.strip().lower()
            if not text:
                continue
            for marker in _SIGNATURE_MARKERS:
                if marker in text:
                    return True
    return False


def _to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_amount(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    candidate = text.replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(candidate)
    except ValueError:
        return None


def _to_date(value: Any) -> Date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, Date):
        return value
    if isinstance(value, (int, float)):
        # xlrd already converts dates upstream; a bare number here is unusual.
        try:
            return xlrd.xldate.xldate_as_datetime(float(value), 0).date()
        except (xlrd.XLDateError, ValueError):
            return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None
